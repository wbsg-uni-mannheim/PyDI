#!/usr/bin/env python3
"""Build per-domain XLSX reports under ``usecases_synthetic/statistics/``.

Central reporting for the variant pipeline. Each domain gets one XLSX
file. The first sheet (``evaluation_legend``) explains the per-stage
evaluation surfaces; the rest are data.

Sheets, in order:

1. ``evaluation_legend`` — glossary explaining (a) why EM matching has
   four separate evaluation surfaces under the R7b dual-model dual-test
   infrastructure and (b) which committee_summary row corresponds to
   which surface. Read first if any metric name in the file is unclear.
2. ``sizes`` — source row counts across baseline + every level
   (easy / medium / hard) that has been generated.
3. ``splits`` — combined sheet covering (a) EM gold per-pair × per-split
   (train / val / test) ``total`` / ``positive`` / ``negative`` /
   ``pos_rate`` and (b) fusion validation + test entity counts.
   Variants use the regenerated copies of EM gold when present.
4. ``examples`` — sampled positive + negative gold pairs from each EM
   pair, rendered side-by-side across baseline / easy / medium / hard
   so surface drift (K1/K5/K6/K8/K10) is visible per pair.
5. ``transformations`` — per-cluster per-field value drift across
   levels, lining up baseline / easy / medium / hard for the same 10
   clusters as the ``examples`` sheet.
6. ``committee_summary`` — per-stage committee ``macro`` headline + the
   best individual member's score on each level's primary surface.
   For EM matching this row is the ``train=Variant test=Variant``
   surface; the other three EM matching surfaces appear in dedicated
   sheets below.
7. ``per_member`` — every committee member's primary headline metric
   for every stage × level cell.
8. ``selection_map`` — per-attribute member selection for C12
   optimised members (``pydi_per_attribute_optimal``,
   ``rule_per_attribute_optimal``).
9. ``EM match (train=BL test=BL)`` — EM matching, baseline-trained
   model evaluated on the baseline test gold pairs.
10. ``EM match (train=BL test=Var)`` — EM matching, baseline-trained
    model evaluated on the variant's K2-regenerated test gold.
11. ``EM match (train=Var test=BL)`` — EM matching, variant-trained
    model evaluated on the baseline test gold.
12. ``EM match (train=Var test=Var)`` — EM matching, variant-trained
    model evaluated on the variant's regenerated test gold (paper
    headline; matches the committee_summary row).
13. ``EM block (train=BL test=BL)`` through ``EM block (train=Var
    test=Var)`` — same 4-surface layout for EM blocking. Runner-side
    dual-test wiring for blockers is an R7c follow-up; until it lands
    these cells are populated as 0.0 placeholders. The legend sheet
    documents the runner state.

Run::

    python usecases_synthetic/scripts/build_statistics.py
    python usecases_synthetic/scripts/build_statistics.py --domain music
    python usecases_synthetic/scripts/build_statistics.py --domain music --domain games

Outputs land at ``usecases_synthetic/statistics/<domain>.xlsx``. The
script is idempotent — re-run after each new S.7 / sanity-ladder
iteration to refresh the spreadsheets.

Where the underlying data lives:

- Baseline metrics: ``usecases_synthetic/baselines/<domain>/baseline_metrics.json``
- Level metrics: ``usecases_synthetic/validation/<domain>/<level>/metrics.json``
- Baseline sources / EM gold / fusion gold: ``usecases/<domain>/input/``
  (or ``usecases_synthetic/usecases/<domain>/input/`` when the domain
  config sets ``data_root``).
- Variant sources / EM gold / fusion gold:
  ``usecases/<domain>-augmented/<level>/input/`` (likewise honoring the
  data_root override).

Outputs reference [scripts/build_statistics.py](.) — see also the
``Central reporting`` note in [plans/plan_s1_final.md](../../plans/plan_s1_final.md).
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from pathlib import Path
from typing import Any, Mapping
from xml.etree import ElementTree as ET

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from usecases_synthetic.lib.domain_config import (  # noqa: E402
    USECASES_DIR,
    load_domain_config,
)

logger = logging.getLogger(__name__)

LEVELS: tuple[str, ...] = ("baseline", "easy", "medium", "hard")
DEFAULT_DOMAINS: tuple[str, ...] = ("music", "games", "products", "companies")

# Per-stage headline metric (committee aggregated).
#
# R7b dual-model dual-test (2026-05-27): EM matching + blocking
# headlines switch from the baseline-model-on-regen alias keys to the
# explicit variant-model-on-regen keys. Output of ``aggregated`` from
# the committee runners has the new keys; legacy outputs predating
# R7b carry only the alias — readers should fall back via
# ``dict.get(new, dict.get(legacy, 0.0))``.
_STAGE_AGG_KEY: dict[str, str] = {
    "sm": "macro_f1",
    "norm": "macro_f1",
    # EM blocking is unsupervised for every member except sc_block, so
    # the R7b dual-test split (variant_model vs baseline_model) is a
    # no-op for blocking — the runner still emits the keys but writes
    # 0.0 to all four. Use the populated single-test metric instead.
    "em_blocking": "macro_pair_recall",
    "em_matching": "macro_f1_variant_model_on_regen_test",
    "fusion": "overall_accuracy",
}

# Per-stage headline metric (per-member). For EM matching/blocking the
# headline member metric is ``f1`` / ``pair_recall`` — these are the
# legacy keys that the per-pair dict still emits via the committee's
# fallback chain (which picks variant_model_on_regen_test when the
# variant model is distinct from baseline; see plan_revision.md R7b).
_STAGE_MEMBER_KEY: dict[str, str] = {
    "sm": "f1",
    "norm": "macro_f1",
    "em_blocking": "pair_recall",
    "em_matching": "f1",
    "fusion": "macro_accuracy",
}

# R7b dual-model dual-test cross-product (2026-05-27). Per-stage, each
# member is evaluated on {baseline-trained, variant-trained} model x
# {baseline-test, variant-regenerated-test} = 4 surfaces. The runner
# writes all four into ``aggregated`` + ``per_member`` of the stage
# block. The four surfaces get dedicated sheets per stage so a reader
# can compare them side by side without guessing which metric the
# committee_summary row shows.
#
# Stages with dual-test surfaces:
#   - em_matching: all four populated by the runner (R7b live).
#   - em_blocking: dual-test keys emitted but populated as 0.0; the
#     runner-side wiring for blocking dual-test is an R7c follow-up
#     (see plan_revision.md R7c). The sheets are still emitted so they
#     are forward-compatible — values fill in automatically once the
#     runner produces them.
#
# Other stages (SM, Norm, Fusion) have a single evaluation surface
# per level and do not appear here.
_DUAL_TEST_SUFFIXES: tuple[tuple[str, str, str], ...] = (
    # (train_label, test_label, key_suffix)
    ("BL", "BL", "baseline_model_on_baseline_test"),
    ("BL", "Var", "baseline_model_on_regen_test"),
    ("Var", "BL", "variant_model_on_baseline_test"),
    ("Var", "Var", "variant_model_on_regen_test"),
)


_SURFACE_DESCRIPTIONS: dict[tuple[str, str], str] = {
    ("BL", "BL"): (
        "Member fit on the BASELINE (un-perturbed) training data and "
        "evaluated against the BASELINE test gold. Reference "
        "evaluation: no variant influence on either side. At the "
        "baseline level this equals the legacy single-test metric."
    ),
    ("BL", "Var"): (
        "Member fit on baseline training data, evaluated against the "
        "VARIANT's K2-regenerated test gold. 'How well does a clean-"
        "data model handle a perturbed test set?' Gap vs the BL/BL "
        "surface isolates test-side difficulty."
    ),
    ("Var", "BL"): (
        "Member fit on the VARIANT's perturbed training data, "
        "evaluated against the original BASELINE test gold. Measures "
        "whether variant-aware training degrades clean-data "
        "performance. Gap vs the BL/BL surface isolates train-side "
        "robustness cost."
    ),
    ("Var", "Var"): (
        "Member fit on VARIANT training data, evaluated against the "
        "VARIANT's regenerated test gold. Both train and test see the "
        "same perturbation distribution. Primary headline metric for "
        "the paper; matches the corresponding row in committee_summary."
    ),
}


def _build_surface_spec(
    stage: str,
    label_short: str,
    agg_prefix: str,
    member_prefix: str,
    runner_note: str = "",
) -> tuple[dict[str, str], ...]:
    """Build the 4-surface dual-test spec list for one stage.

    Parameters
    ----------
    stage : str
        Stage identifier (``em_matching`` / ``em_blocking``).
    label_short : str
        Short label used in sheet names (``EM match`` / ``EM block``).
    agg_prefix : str
        Aggregated-metric prefix (``macro_f1`` / ``macro_pair_recall``).
    member_prefix : str
        Per-member metric prefix (``f1`` / ``pair_recall``).
    runner_note : str, optional
        Extra paragraph appended to each surface description (used to
        flag em_blocking's current 0.0-placeholder state).
    """
    surfaces: list[dict[str, str]] = []
    for train_label, test_label, key_suffix in _DUAL_TEST_SUFFIXES:
        agg_key = f"{agg_prefix}_{key_suffix}"
        member_key = f"{member_prefix}_{key_suffix}"
        sheet_name = f"{label_short} (train={train_label} test={test_label})"
        title = (
            f"{stage.replace('_', ' ')} - "
            f"{train_label}-trained model on {test_label}-test gold"
        )
        description = _SURFACE_DESCRIPTIONS[(train_label, test_label)]
        if runner_note:
            description = description + "\n\n" + runner_note
        surfaces.append(
            {
                "stage": stage,
                "agg_key": agg_key,
                "member_key": member_key,
                "sheet_name": sheet_name,
                "title": title,
                "description": description,
            }
        )
    return tuple(surfaces)


_EM_MATCHING_SURFACES: tuple[dict[str, str], ...] = _build_surface_spec(
    stage="em_matching",
    label_short="EM match",
    agg_prefix="macro_f1",
    member_prefix="f1",
)


_EM_BLOCKING_RUNNER_NOTE = (
    "Runner state (2026-05-28): the em_blocking dual-test wiring "
    "is not yet implemented (R7c follow-up). The runner emits the "
    "four dual-test keys as 0.0 placeholders. The committee row in "
    "the committee_summary sheet uses the single-test "
    "macro_pair_recall — that is the only meaningful blocking "
    "metric today. These sheets become populated automatically "
    "once R7c lands."
)


_EM_BLOCKING_SURFACES: tuple[dict[str, str], ...] = _build_surface_spec(
    stage="em_blocking",
    label_short="EM block",
    agg_prefix="macro_pair_recall",
    member_prefix="pair_recall",
    runner_note=_EM_BLOCKING_RUNNER_NOTE,
)


_ALL_DUAL_TEST_SURFACES: tuple[dict[str, str], ...] = (
    _EM_MATCHING_SURFACES + _EM_BLOCKING_SURFACES
)

STATISTICS_DIR: Path = REPO_ROOT / "usecases_synthetic" / "statistics"


def _data_root_for_domain(domain: str) -> Path:
    """Resolve the ``usecases/`` root honoring ``data_root`` overrides.

    Products lives under ``usecases_synthetic/usecases/products/`` via
    the ``data_root: usecases_synthetic/usecases`` override; other
    domains use the canonical ``usecases/`` directly.
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        return USECASES_DIR
    if cfg.data_root is not None:
        root = REPO_ROOT / cfg.data_root
        if root.exists():
            return root
    return USECASES_DIR


def _csv_rowcount(path: Path) -> int | None:
    """Count data rows in a CSV (excluding header if present).

    Conservative header detection: if the first line contains any
    non-id-looking column name, treat it as a header.
    """
    if not path.exists():
        return None
    try:
        with open(path, encoding="utf-8") as f:
            lines = sum(1 for _ in f)
    except (OSError, UnicodeDecodeError):
        return None
    if lines == 0:
        return 0
    # EM gold files are headerless (`id1,id2,label`); the runner detects
    # this and re-reads. For our row-count purposes, treat as +1 row
    # for files that look like data without header. We can't perfectly
    # distinguish; subtract 1 conservatively when first cell looks like
    # a column name, else return raw.
    try:
        with open(path, encoding="utf-8") as f:
            first = f.readline().strip()
    except (OSError, UnicodeDecodeError):
        return lines
    first_cell = first.split(",")[0]
    # EM gold convention: id-like first cell ("dbpedia_123", "mbrainz_4")
    # vs header convention: column name ("id1", "id", etc.).
    if first_cell in {"id1", "id", "id_left", "source_1"} or first_cell.startswith(
        ("Attribute_",)
    ):
        return max(0, lines - 1)
    # Heuristic: if first cell contains an underscore and a digit it's
    # likely a record id, no header.
    if any(c.isdigit() for c in first_cell) and "_" in first_cell:
        return lines
    # Otherwise assume header.
    return max(0, lines - 1)


_TRUE_LABELS = {"true", "1", "yes", "match"}
_FALSE_LABELS = {"false", "0", "no", "nonmatch", "non-match"}


def _em_split_counts(path: Path) -> tuple[int, int, int] | None:
    """Parse an EM gold CSV; return ``(total, positive, negative)``.

    Handles both supported formats:
    - Headerless: ``id1,id2,label`` (label = TRUE / FALSE / true / false)
    - Header'd:   ``id1,id2,source_1,source_2,label`` (regenerated files)

    Returns ``None`` on parse failure. Rows with unrecognised labels are
    counted toward ``total`` but neither positive nor negative — those
    should not appear in well-formed gold but we don't want to mask a
    bad file with a silent skip.
    """
    if not path.exists():
        return None
    total = 0
    positive = 0
    negative = 0
    try:
        with open(path, encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            try:
                first = next(reader)
            except StopIteration:
                return (0, 0, 0)
            # Header detection: a row whose first cell is a known column
            # name (id1/id) is a header; otherwise treat as data.
            has_header = first and first[0].strip().lower() in {
                "id1",
                "id",
                "id_left",
                "source_1",
            }
            if not has_header:
                # Re-process ``first`` as data.
                rows: Any = [first]
                # Chain remaining rows from the reader.

                def _iter() -> Any:
                    yield from rows
                    yield from reader

                source = _iter()
            else:
                source = reader
            for row in source:
                if not row:
                    continue
                total += 1
                label_cell = row[-1].strip().lower()
                if label_cell in _TRUE_LABELS:
                    positive += 1
                elif label_cell in _FALSE_LABELS:
                    negative += 1
    except (OSError, UnicodeDecodeError, csv.Error):
        return None
    return (total, positive, negative)


def _xml_entity_count(path: Path) -> int | None:
    """Count top-level entity elements in a fusion XML file."""
    if not path.exists():
        return None
    try:
        tree = ET.parse(path)
    except ET.ParseError:
        return None
    root = tree.getroot()
    # The fusion XML root has children = entities.
    return len(list(root))


def _resolve_variant_root(domain: str, level: str) -> Path:
    """Return the ``input/`` root for a (domain, level) pair.

    Mirrors ``variant_loader._variant_root``: the per-domain ``data_root``
    override applies **only to baseline** (e.g. products' baseline lives
    under ``usecases_synthetic/usecases/products/``). Variants always
    live under the canonical top-level ``usecases/<domain>-augmented/<level>/``
    for cross-domain consistency.
    """
    if level == "baseline":
        return _data_root_for_domain(domain) / domain
    return USECASES_DIR / f"{domain}-augmented" / level


def _collect_sizes(domain: str) -> dict[str, dict[str, int | None]]:
    """Build ``{label: {level: count}}`` for the sizes sheet.

    Labels cover each source ``<source>_rows``. EM + fusion splits live
    on the dedicated ``splits`` sheet via :func:`_collect_em_splits` /
    :func:`_collect_fusion_splits`. Missing files / levels emit ``None``
    so the spreadsheet shows blanks instead of zeros.
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        logger.warning("Domain config missing for %s; skipping sizes", domain)
        return {}

    sources = [src.name for src in cfg.sources]

    sizes: dict[str, dict[str, int | None]] = {}
    for level in LEVELS:
        variant_root = _resolve_variant_root(domain, level)
        data_dir = variant_root / "input" / "data"

        # Per-source row counts.
        for src in sources:
            label = f"source_{src}_rows"
            for ext in (".csv", ".json"):
                p = data_dir / f"{src}{ext}"
                if p.exists():
                    if ext == ".csv":
                        count = _csv_rowcount(p)
                    else:
                        # JSON: count records by line if JSONL, else
                        # try to parse and count elements.
                        try:
                            with open(p, encoding="utf-8") as f:
                                head = f.read(2)
                            if head.startswith("["):
                                count = sum(
                                    1 for _ in json.load(open(p, encoding="utf-8"))
                                )
                            else:
                                count = sum(1 for _ in open(p, encoding="utf-8"))
                        except (OSError, json.JSONDecodeError):
                            count = None
                    sizes.setdefault(label, {})[level] = count
                    break
            else:
                sizes.setdefault(label, {})[level] = None

    return sizes


def _collect_fusion_splits(
    domain: str,
) -> dict[str, dict[str, int | None]]:
    """Return fusion validation/test entity counts per level.

    Keys are ``"validation"`` and ``"test"``; each maps to
    ``{level: count}``. ``None`` indicates the XML file is absent at
    that level.
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        return {}
    out: dict[str, dict[str, int | None]] = {"validation": {}, "test": {}}
    fusion_files = getattr(cfg, "fusion_files", None)
    for level in LEVELS:
        fusion_dir = _resolve_variant_root(domain, level) / "input" / "fusion"
        for split_key, attr in (("validation", "validation"), ("test", "test")):
            fname = None
            if fusion_files is not None and isinstance(fusion_files, Mapping):
                fname = fusion_files.get(attr)
            if not fname:
                fname = f"{attr}_set.xml" if attr == "validation" else "test_set.xml"
            count: int | None = None
            for candidate in (fname, f"{attr}_set.xml", f"{attr}_set_final.xml"):
                p = fusion_dir / candidate
                if p.exists():
                    count = _xml_entity_count(p)
                    break
            out[split_key][level] = count
    return out


def _collect_em_splits(
    domain: str,
) -> tuple[
    list[tuple[str, str]],
    dict[tuple[str, str, str], dict[str, int | None]],
]:
    """Build the per-(pair, split) total/positive/negative table.

    Returns ``(pairs, table)`` where ``pairs`` is the ordered list of
    ``(pair_name, split)`` tuples (e.g. ``("musicbrainz_2_discogs",
    "train")``) and ``table`` maps ``(pair_name, split, metric)`` to
    ``{level: value}``. ``metric`` is one of ``total`` / ``positive`` /
    ``negative`` / ``pos_rate``.

    For variants the regenerated copy of the gold is used when present;
    falls back to the original. Reverse-name files are also tried so
    canonical pairs authored in opposing orientations (e.g. games'
    ``(metacritic, dbpedia)`` whose test is ``dbpedia_2_metacritic_test.csv``)
    still surface.
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        return ([], {})

    pairs = [tuple(p) for p in cfg.source_pairs]
    splits = ("train", "val", "test")

    ordered: list[tuple[str, str]] = []
    table: dict[tuple[str, str, str], dict[str, int | None]] = {}
    for src1, src2 in pairs:
        pair_name = f"{src1}_2_{src2}"
        reverse_name = f"{src2}_2_{src1}"
        for split in splits:
            ordered.append((pair_name, split))
            for level in LEVELS:
                variant_root = _resolve_variant_root(domain, level)
                em_dir = variant_root / "input" / "entitymatching"
                if level == "baseline":
                    candidates = [
                        em_dir / f"{pair_name}_{split}.csv",
                        em_dir / f"{reverse_name}_{split}.csv",
                    ]
                else:
                    candidates = [
                        em_dir / f"{pair_name}_{split}_regenerated.csv",
                        em_dir / f"{reverse_name}_{split}_regenerated.csv",
                        em_dir / f"{pair_name}_{split}.csv",
                        em_dir / f"{reverse_name}_{split}.csv",
                    ]
                counts: tuple[int, int, int] | None = None
                for p in candidates:
                    if p.exists():
                        counts = _em_split_counts(p)
                        break
                total_v: int | None
                pos_v: int | None
                neg_v: int | None
                rate_v: float | None
                if counts is None:
                    total_v = pos_v = neg_v = None
                    rate_v = None
                else:
                    total_v, pos_v, neg_v = counts
                    rate_v = (pos_v / total_v) if total_v > 0 else None
                table.setdefault((pair_name, split, "total"), {})[level] = total_v
                table.setdefault((pair_name, split, "positive"), {})[level] = pos_v
                table.setdefault((pair_name, split, "negative"), {})[level] = neg_v
                # ``pos_rate`` carries a float; store as ``Any`` via the
                # same dict to keep the writer simple.
                table.setdefault((pair_name, split, "pos_rate"), {})[
                    level
                ] = rate_v  # type: ignore[assignment]
    return (ordered, table)


EXAMPLES_CLUSTER_COUNT: int = 10
EXAMPLES_RECORD_CHAR_CAP: int = 600


def _load_source_records(
    domain: str,
    level: str,
    source_name: str,
    id_prefix: str,
    id_column_hint: str | None,
) -> dict[str, dict[str, str]]:
    """Load a source CSV/JSON and return ``{id: {col: value, ...}}``.

    Tries ``<source>.csv`` first, then ``<source>.json`` (array form).
    Products' baseline ships as JSON while variants land as CSV, so the
    extension is per-level. The synthetic loader renames ``id_column``
    → ``id`` for variants, so non-baseline files generally just use
    ``id``; baseline may use a configured ``id_column`` or prefix-match.
    """
    variant_root = _resolve_variant_root(domain, level)
    base = variant_root / "input" / "data" / source_name
    try:
        import pandas as pd
    except ImportError:
        return {}

    df = None
    csv_path = base.with_suffix(".csv")
    json_path = base.with_suffix(".json")
    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
        except Exception:  # noqa: BLE001 — best-effort sampling
            df = None
    if df is None and json_path.exists():
        try:
            with open(json_path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                df = pd.DataFrame.from_records(data).astype(str).fillna("")
        except (OSError, json.JSONDecodeError, ValueError):
            df = None
    if df is None or df.empty:
        return {}

    # Identify the id column. Preference order: ``id``, then the
    # configured ``id_column``, then any column whose values look like
    # ``<id_prefix><digits>``.
    id_col: str | None = None
    if "id" in df.columns:
        id_col = "id"
    elif id_column_hint and id_column_hint in df.columns:
        id_col = id_column_hint
    else:
        for c in df.columns:
            sample = df[c].astype(str).head(20)
            if any(v.startswith(id_prefix) for v in sample):
                id_col = c
                break
    if id_col is None:
        return {}

    records: dict[str, dict[str, str]] = {}
    for _, row in df.iterrows():
        rid = str(row[id_col])
        if not rid:
            continue
        record = {c: str(row[c]) for c in df.columns if c != id_col}
        records[rid] = record
    return records


def _format_record(record: dict[str, str] | None) -> str:
    """Render a record dict as ``k1=v1; k2=v2; ...``.

    Skips empty values. Truncates the joined string at
    ``EXAMPLES_RECORD_CHAR_CAP`` (with an ellipsis) so Excel renders it
    cleanly. ``None`` means the record was not found in the level's
    source CSV (K3 drop).
    """
    if record is None:
        return "<dropped>"
    parts: list[str] = []
    for k, v in record.items():
        if v == "" or v.lower() in {"nan", "none"}:
            continue
        parts.append(f"{k}={v}")
    s = "; ".join(parts)
    if len(s) > EXAMPLES_RECORD_CHAR_CAP:
        s = s[: EXAMPLES_RECORD_CHAR_CAP - 1] + "…"
    return s


def _read_em_gold_pairs(
    em_dir: Path, pair_name: str, reverse_name: str, split: str, regenerated: bool
) -> list[tuple[str, str, bool]] | None:
    """Return ``[(id1, id2, is_positive), ...]`` from an EM gold CSV.

    Handles both headerless and header'd formats and reverse-name
    fallback. Returns ``None`` when no file is found.
    """
    if regenerated:
        candidates = [
            em_dir / f"{pair_name}_{split}_regenerated.csv",
            em_dir / f"{reverse_name}_{split}_regenerated.csv",
        ]
    else:
        candidates = [
            em_dir / f"{pair_name}_{split}.csv",
            em_dir / f"{reverse_name}_{split}.csv",
        ]
    for p in candidates:
        if not p.exists():
            continue
        out: list[tuple[str, str, bool]] = []
        try:
            with open(p, encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                try:
                    first = next(reader)
                except StopIteration:
                    return []
                has_header = first and first[0].strip().lower() in {
                    "id1",
                    "id",
                    "id_left",
                    "source_1",
                }
                # If reverse-name file matched, swap id1/id2 so they
                # always come back in (src1, src2) order.
                swap = p.name.startswith(reverse_name)
                if not has_header:
                    rows: list[list[str]] = [first]

                    def _iter() -> Any:
                        yield from rows
                        yield from reader

                    source = _iter()
                else:
                    source = reader
                for row in source:
                    if len(row) < 3:
                        continue
                    a, b = row[0], row[1]
                    label_cell = row[-1].strip().lower()
                    is_pos = label_cell in _TRUE_LABELS
                    is_neg = label_cell in _FALSE_LABELS
                    if not (is_pos or is_neg):
                        continue
                    if swap:
                        a, b = b, a
                    out.append((a, b, is_pos))
        except (OSError, UnicodeDecodeError, csv.Error):
            return None
        return out
    return None


def _record_value_set(record: dict[str, str] | None) -> set[str]:
    """Set of non-trivial string values in a record, ignoring column names.

    Used to score baseline-vs-hard drift while staying invariant to K8
    column renames — only actual value changes contribute.
    """
    if record is None:
        return set()
    out: set[str] = set()
    for v in record.values():
        s = v.strip()
        if not s:
            continue
        if s.lower() in {"nan", "none", "null"}:
            continue
        out.add(s)
    return out


def _value_drift(baseline: dict[str, str] | None, hard: dict[str, str] | None) -> float:
    """Jaccard distance between baseline and hard value sets.

    Returns 1.0 when one side is missing (K3 drop) or the value sets are
    disjoint, 0.0 when identical. Invariant to column renames.
    """
    if baseline is None and hard is None:
        return 0.0
    if baseline is None or hard is None:
        return 1.0
    a = _record_value_set(baseline)
    b = _record_value_set(hard)
    if not a and not b:
        return 0.0
    union = a | b
    if not union:
        return 0.0
    return len(a ^ b) / len(union)


def _build_baseline_clusters(domain: str, cfg: Any) -> list[dict[str, str]]:
    """Union-find over baseline gold positives → cluster member sets.

    Returns a list of ``{member_id: source_name}`` dicts, one per
    connected component spanning ≥2 distinct sources. Negatives are
    ignored; only positive matches link records into clusters.
    """
    pairs = [tuple(p) for p in cfg.source_pairs]
    baseline_dir = (
        _resolve_variant_root(domain, "baseline") / "input" / "entitymatching"
    )

    parent: dict[str, str] = {}

    def find(x: str) -> str:
        while parent.get(x, x) != x:
            parent[x] = parent.get(parent[x], parent[x])
            x = parent[x]
        return x

    def union(x: str, y: str) -> None:
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[rx] = ry

    id_to_source: dict[str, str] = {}
    for src1, src2 in pairs:
        pair_name = f"{src1}_2_{src2}"
        reverse_name = f"{src2}_2_{src1}"
        for split in ("train", "val", "test"):
            gold = _read_em_gold_pairs(
                baseline_dir, pair_name, reverse_name, split, regenerated=False
            )
            if not gold:
                continue
            for id1, id2, is_pos in gold:
                if not is_pos:
                    continue
                parent.setdefault(id1, id1)
                parent.setdefault(id2, id2)
                id_to_source.setdefault(id1, src1)
                id_to_source.setdefault(id2, src2)
                union(id1, id2)

    components: dict[str, dict[str, str]] = {}
    for x in parent:
        components.setdefault(find(x), {})[x] = id_to_source[x]

    return [
        members
        for members in components.values()
        if len({src for src in members.values()}) >= 2
    ]


def _collect_corner_examples(domain: str) -> list[dict[str, Any]]:
    """Pick 10 high-drift entity clusters and render their pair edges.

    Selection: build clusters via union-find over baseline gold
    positives, score each cluster by sum of ``Jaccard(baseline_values,
    hard_values)`` across its members (K8 column renames contribute
    zero — only actual value changes matter), then take the top-10 with
    valid baseline records and ≥2 distinct sources.

    Each cluster expands to one row per ``(source_pair, level)`` where
    both member IDs exist in baseline. Returns dicts shaped for
    :func:`_write_examples_sheet`:

    ``{"cluster_id", "cluster_label", "pair", "id1", "id2",
       "records": {level: (left_str, right_str)}}``
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        return []

    src_specs = {s.name: s for s in cfg.sources}
    pairs = [tuple(p) for p in cfg.source_pairs]
    if not pairs:
        return []

    # Per-level per-source record cache.
    record_cache: dict[tuple[str, str], dict[str, dict[str, str]]] = {}

    def _records(level: str, src_name: str) -> dict[str, dict[str, str]]:
        key = (level, src_name)
        if key not in record_cache:
            spec = src_specs.get(src_name)
            id_prefix = spec.id_prefix if spec else ""
            id_col_hint = spec.id_column if spec else None
            record_cache[key] = _load_source_records(
                domain, level, src_name, id_prefix, id_col_hint
            )
        return record_cache[key]

    clusters = _build_baseline_clusters(domain, cfg)
    if not clusters:
        return []

    scored: list[tuple[float, dict[str, str]]] = []
    for members in clusters:
        # Require every member to exist at *every* level so the rendered
        # rows never contain ``<dropped>`` placeholders. K2 entity drops
        # at variant levels would otherwise show up as ``<dropped>`` in
        # the examples sheet and obscure the perturbation we want to
        # showcase. Drift is still scored against the hard level.
        valid = True
        score = 0.0
        for mid, src in members.items():
            present_at_every_level = all(
                _records(level, src).get(mid) is not None for level in LEVELS
            )
            if not present_at_every_level:
                valid = False
                break
            bl = _records("baseline", src).get(mid)
            hr = _records("hard", src).get(mid)
            score += _value_drift(bl, hr)
        if not valid or score == 0.0:
            continue
        scored.append((score, members))

    # Sort by drift desc; secondary key = smaller cluster (more legible).
    scored.sort(key=lambda x: (-x[0], len(x[1])))
    picked = scored[:EXAMPLES_CLUSTER_COUNT]

    examples: list[dict[str, Any]] = []
    for idx, (_score, members) in enumerate(picked, start=1):
        # Build a stable label listing members in source order.
        by_source: dict[str, list[str]] = {}
        for mid, src in members.items():
            by_source.setdefault(src, []).append(mid)
        ordered_members: list[str] = []
        for src in [s.name for s in cfg.sources]:
            for mid in sorted(by_source.get(src, [])):
                ordered_members.append(f"{mid}")
        cluster_label = f"Cluster {idx}: " + " | ".join(ordered_members)

        # For each configured source pair, find a (left, right) inside
        # the cluster (both must be members). If multiple candidates,
        # take the first deterministically.
        for src1, src2 in pairs:
            pair_name = f"{src1}_2_{src2}"
            lefts = sorted(by_source.get(src1, []))
            rights = sorted(by_source.get(src2, []))
            if not lefts or not rights:
                continue
            id1, id2 = lefts[0], rights[0]
            per_level: dict[str, tuple[str, str]] = {}
            for level in LEVELS:
                left = _records(level, src1).get(id1)
                right = _records(level, src2).get(id2)
                per_level[level] = (_format_record(left), _format_record(right))
            examples.append(
                {
                    "cluster_id": idx,
                    "cluster_label": cluster_label,
                    "pair": pair_name,
                    "id1": id1,
                    "id2": id2,
                    "records": per_level,
                }
            )
    return examples


def _collect_record_transformations(domain: str) -> list[dict[str, Any]]:
    """Build per-record per-field transformation tables for the top clusters.

    Uses the same cluster pick as :func:`_collect_corner_examples` so
    the ``examples`` and ``transformations`` sheets cross-reference.
    For each cluster member, emits a block showing every field's value
    across baseline / easy / medium / hard.

    Field alignment is position-based against the baseline column order
    — K8 only renames columns (never reorders / drops them), so the
    canonical baseline column name labels each row and the corresponding
    value at each level comes from the same positional slot. The picker
    requires every member to survive at every level (matches the
    examples-sheet rule), so ``<dropped>`` would normally not appear;
    the placeholder is kept as a defensive write for the edge case where
    a record disappears mid-load (e.g. a partially written variant
    directory).

    Returns a list of dicts with keys ``cluster_id`` / ``cluster_label``
    / ``record_id`` / ``source`` / ``fields``, where ``fields`` is a
    list of ``(baseline_name, {level: value})`` tuples.
    """
    try:
        cfg = load_domain_config(domain)
    except FileNotFoundError:
        return []

    src_specs = {s.name: s for s in cfg.sources}
    if not cfg.source_pairs:
        return []

    record_cache: dict[tuple[str, str], dict[str, dict[str, str]]] = {}

    def _records(level: str, src_name: str) -> dict[str, dict[str, str]]:
        key = (level, src_name)
        if key not in record_cache:
            spec = src_specs.get(src_name)
            id_prefix = spec.id_prefix if spec else ""
            id_col_hint = spec.id_column if spec else None
            record_cache[key] = _load_source_records(
                domain, level, src_name, id_prefix, id_col_hint
            )
        return record_cache[key]

    def _columns(level: str, src_name: str) -> list[str]:
        """Return non-id column order for the given (level, source).

        Derived from the first available record dict (all records in a
        given file share the same keys / order).
        """
        recs = _records(level, src_name)
        if not recs:
            return []
        first = next(iter(recs.values()))
        return list(first.keys())

    clusters = _build_baseline_clusters(domain, cfg)
    if not clusters:
        return []

    # Re-run the same scoring + pick as the examples sheet so both
    # sheets feature the same 10 clusters. Selection requires every
    # member to exist at every level (no ``<dropped>`` rows; see
    # :func:`_collect_corner_examples`).
    scored: list[tuple[float, dict[str, str]]] = []
    for members in clusters:
        valid = True
        score = 0.0
        for mid, src in members.items():
            present_at_every_level = all(
                _records(level, src).get(mid) is not None for level in LEVELS
            )
            if not present_at_every_level:
                valid = False
                break
            bl = _records("baseline", src).get(mid)
            hr = _records("hard", src).get(mid)
            score += _value_drift(bl, hr)
        if not valid or score == 0.0:
            continue
        scored.append((score, members))
    scored.sort(key=lambda x: (-x[0], len(x[1])))
    picked = scored[:EXAMPLES_CLUSTER_COUNT]

    out: list[dict[str, Any]] = []
    source_order = [s.name for s in cfg.sources]
    for idx, (_score, members) in enumerate(picked, start=1):
        by_source: dict[str, list[str]] = {}
        for mid, src in members.items():
            by_source.setdefault(src, []).append(mid)
        ordered_members: list[str] = []
        for src in source_order:
            for mid in sorted(by_source.get(src, [])):
                ordered_members.append(mid)
        cluster_label = f"Cluster {idx}: " + " | ".join(ordered_members)

        # Emit one block per cluster member, in source order.
        for src in source_order:
            for mid in sorted(by_source.get(src, [])):
                baseline_cols = _columns("baseline", src)
                if not baseline_cols:
                    continue
                fields: list[tuple[str, dict[str, str]]] = []
                for col_idx, baseline_name in enumerate(baseline_cols):
                    per_level_values: dict[str, str] = {}
                    for level in LEVELS:
                        rec = _records(level, src).get(mid)
                        cols = _columns(level, src)
                        if rec is None:
                            per_level_values[level] = "<dropped>"
                        elif col_idx < len(cols):
                            level_col = cols[col_idx]
                            per_level_values[level] = rec.get(level_col, "")
                        else:
                            per_level_values[level] = ""
                    fields.append((baseline_name, per_level_values))
                out.append(
                    {
                        "cluster_id": idx,
                        "cluster_label": cluster_label,
                        "record_id": mid,
                        "source": src,
                        "fields": fields,
                    }
                )
    return out


def _stage_agg_value(stage_block: Mapping[str, Any], stage: str) -> float | None:
    """Return the stage's committee aggregated headline metric.

    R7b: for EM stages, prefer the new variant-model-on-regen-test key
    but fall back to the pre-R7b alias when reading legacy outputs.
    """
    key = _STAGE_AGG_KEY.get(stage)
    if key is None:
        return None
    agg = stage_block.get("aggregated") or {}
    if not isinstance(agg, Mapping):
        return None
    val = agg.get(key)
    if val is None and stage == "em_matching":
        val = agg.get("macro_f1_regen_test")
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _stage_best_member(
    stage_block: Mapping[str, Any], stage: str
) -> tuple[str, float | None]:
    """Return ``(member_name, value)`` for the highest-scoring member."""
    member_key = _STAGE_MEMBER_KEY.get(stage)
    if member_key is None:
        return ("", None)
    per_member = stage_block.get("per_member") or {}
    best_name = ""
    best_val: float | None = None
    for name, body in per_member.items():
        if not isinstance(body, Mapping):
            continue
        metrics = body.get("metrics") or {}
        if not isinstance(metrics, Mapping):
            continue
        v = metrics.get(member_key)
        try:
            v_f = float(v) if v is not None else None
        except (TypeError, ValueError):
            v_f = None
        if v_f is None:
            continue
        if best_val is None or v_f > best_val:
            best_val = v_f
            best_name = name
    return (best_name, best_val)


def _all_members_for_stage(
    per_level_metrics: Mapping[str, Mapping[str, Any]], stage: str
) -> list[str]:
    """Union of member names appearing in any level for ``stage``."""
    names: set[str] = set()
    for body in per_level_metrics.values():
        per_stage = body.get("per_stage") or {}
        if not isinstance(per_stage, Mapping):
            continue
        block = per_stage.get(stage) or {}
        if not isinstance(block, Mapping):
            continue
        per_member = block.get("per_member") or {}
        if isinstance(per_member, Mapping):
            names.update(per_member.keys())
    return sorted(names)


def _load_level_metrics(domain: str) -> dict[str, dict[str, Any]]:
    """Return ``{level: metrics_dict}`` for whatever levels exist on disk."""
    out: dict[str, dict[str, Any]] = {}
    baseline_path = (
        REPO_ROOT
        / "usecases_synthetic"
        / "baselines"
        / domain
        / "baseline_metrics.json"
    )
    if baseline_path.exists():
        with open(baseline_path, encoding="utf-8") as f:
            out["baseline"] = json.load(f)
    for level in ("easy", "medium", "hard"):
        p = (
            REPO_ROOT
            / "usecases_synthetic"
            / "validation"
            / domain
            / level
            / "metrics.json"
        )
        if p.exists():
            with open(p, encoding="utf-8") as f:
                out[level] = json.load(f)
    return out


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_GROUP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_GROUP_FONT = Font(bold=True)


def _autosize(ws) -> None:
    """Set column widths to roughly fit content.

    Walks rows (not ``ws.columns``) so we never touch MergedCell
    instances; that iteration has been known to confuse Excel readers
    when the sheet contains merged ranges.
    """
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            col_idx = cell.column
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > widths.get(col_idx, 0):
                widths[col_idx] = len(s)
    for col_idx, max_len in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, max_len + 2), 50
        )


def _write_sizes_sheet(wb: Workbook, sizes: dict[str, dict[str, int | None]]) -> None:
    """Emit the size statistics sheet.

    Group headers are full rows with the label in every cell (no merged
    cells). Merged cells + ``column_dimensions`` autosizing have caused
    Excel-side ``needs to repair`` errors in some readers; the un-merged
    rendering keeps the output portable.
    """
    ws = wb.create_sheet("sizes")
    headers = ["measurement"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    row_idx = 2
    # Sources only — EM gold + fusion splits live on the ``splits`` sheet.
    groups: list[tuple[str, list[str]]] = [
        ("Sources (row counts)", sorted(k for k in sizes if k.startswith("source_"))),
    ]
    for group_label, labels in groups:
        if not labels:
            continue
        # Full-width group header row (no merges).
        for col in range(1, len(headers) + 1):
            cell = ws.cell(
                row=row_idx, column=col, value=group_label if col == 1 else ""
            )
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT
        row_idx += 1
        for label in labels:
            ws.cell(row=row_idx, column=1, value=label)
            for col_offset, level in enumerate(LEVELS, start=2):
                v = sizes[label].get(level)
                ws.cell(row=row_idx, column=col_offset, value=v)
            row_idx += 1

    _autosize(ws)


def _write_splits_sheet(
    wb: Workbook,
    em_ordered: list[tuple[str, str]],
    em_table: dict[tuple[str, str, str], dict[str, int | None]],
    fusion: dict[str, dict[str, int | None]],
) -> None:
    """Emit the combined train / val / test sheet (EM + fusion).

    Layout per EM pair: a labelled group-header row, then 12 rows
    (3 splits × 4 metrics: ``total`` / ``positive`` / ``negative`` /
    ``pos_rate``). Fusion gets its own group at the bottom with
    ``validation`` / ``test`` rows × ``entities`` metric.

    Empty cells are meaningful: e.g. games has no val EM files at any
    level, so the val rows render blank across all 4 levels.
    """
    ws = wb.create_sheet("splits")
    headers = ["group", "split", "metric"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2

    def _group_header(label: str) -> None:
        nonlocal row_idx
        # Un-merged labelled row. Avoid leading ``=`` (openpyxl
        # auto-detects formulas → Excel marks workbook corrupt).
        for col in range(1, len(headers) + 1):
            cell = ws.cell(
                row=row_idx, column=col, value=f"[{label}]" if col == 1 else ""
            )
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT
        row_idx += 1

    # EM gold blocks. Splits with no file at any level (e.g. games has
    # no EM val files anywhere) are skipped entirely so the sheet only
    # shows things that actually exist on disk.
    metric_rows = ("total", "positive", "negative", "pos_rate")
    seen_pairs: set[str] = set()
    # Pre-compute which (pair, split) tuples have at least one populated
    # level — used to decide whether to emit the rows AND whether to
    # emit a pair's group header at all.
    populated: set[tuple[str, str]] = set()
    for pair_name, split in em_ordered:
        totals = em_table.get((pair_name, split, "total"), {})
        if any(totals.get(lvl) is not None for lvl in LEVELS):
            populated.add((pair_name, split))
    pairs_with_data = {pair for (pair, _split) in populated}

    for pair_name, split in em_ordered:
        if (pair_name, split) not in populated:
            continue
        if pair_name not in seen_pairs and pair_name in pairs_with_data:
            _group_header(f"EM gold: {pair_name}")
            seen_pairs.add(pair_name)
        for metric in metric_rows:
            ws.cell(row=row_idx, column=1, value=pair_name)
            ws.cell(row=row_idx, column=2, value=split)
            ws.cell(row=row_idx, column=3, value=metric)
            level_map = em_table.get((pair_name, split, metric), {})
            for col_offset, level in enumerate(LEVELS, start=4):
                ws.cell(row=row_idx, column=col_offset, value=level_map.get(level))
            row_idx += 1

    # Fusion block. Skip a split if it has no file at any level; skip
    # the whole group + header if neither split has data.
    fusion_splits_with_data = [
        sk
        for sk in ("validation", "test")
        if any(fusion.get(sk, {}).get(lvl) is not None for lvl in LEVELS)
    ]
    if fusion_splits_with_data:
        _group_header("Fusion (entities)")
        for split_key in fusion_splits_with_data:
            level_map = fusion.get(split_key, {})
            ws.cell(row=row_idx, column=1, value="fusion")
            ws.cell(row=row_idx, column=2, value=split_key)
            ws.cell(row=row_idx, column=3, value="entities")
            for col_offset, level in enumerate(LEVELS, start=4):
                ws.cell(row=row_idx, column=col_offset, value=level_map.get(level))
            row_idx += 1

    # Number-format: pos_rate as 0.0000, integers left raw.
    for r in ws.iter_rows(min_row=2, min_col=4, max_col=len(headers)):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    _autosize(ws)


def _write_examples_sheet(wb: Workbook, examples: list[dict[str, Any]]) -> None:
    """Emit the cluster-examples sheet.

    Examples are grouped by entity cluster (10 clusters picked by max
    value-set drift between baseline and hard — invariant to K8 column
    renames). The picker requires every cluster member to exist at
    every level, so no ``<dropped>`` placeholders appear; surface drift
    from K1/K5/K6 / K10 / K8 is visible reading down a column. Each
    cluster shows every configured source-pair edge that has both
    members; each edge expands to 4 rows (one per level) with the left
    + right records rendered as ``k=v; k=v; ...`` strings.
    """
    ws = wb.create_sheet("examples")
    headers = ["cluster", "pair", "level", "left_id", "right_id", "left", "right"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    # Wrap record cells so multi-field strings stay readable.
    wrap = Alignment(wrap_text=True, vertical="top")
    last_cluster: int | None = None
    for ex in examples:
        if ex["cluster_id"] != last_cluster:
            # Group header per cluster. Avoid leading ``=`` (openpyxl
            # auto-detects formulas).
            header_label = f"[{ex['cluster_label']}]"
            for col in range(1, len(headers) + 1):
                cell = ws.cell(
                    row=row_idx,
                    column=col,
                    value=header_label if col == 1 else "",
                )
                cell.fill = _GROUP_FILL
                cell.font = _GROUP_FONT
            row_idx += 1
            last_cluster = ex["cluster_id"]

        for level in LEVELS:
            left_str, right_str = ex["records"].get(level, ("", ""))
            ws.cell(row=row_idx, column=1, value=ex["cluster_id"])
            ws.cell(row=row_idx, column=2, value=ex["pair"])
            ws.cell(row=row_idx, column=3, value=level)
            ws.cell(row=row_idx, column=4, value=ex["id1"])
            ws.cell(row=row_idx, column=5, value=ex["id2"])
            left_cell = ws.cell(row=row_idx, column=6, value=left_str)
            right_cell = ws.cell(row=row_idx, column=7, value=right_str)
            left_cell.alignment = wrap
            right_cell.alignment = wrap
            row_idx += 1

    _autosize(ws)
    # Force the two record columns wide; autosize caps at 50 which is
    # too narrow for a serialized record.
    ws.column_dimensions[get_column_letter(6)].width = 80
    ws.column_dimensions[get_column_letter(7)].width = 80


def _write_transformations_sheet(
    wb: Workbook, transformations: list[dict[str, Any]]
) -> None:
    """Emit the per-record per-field transformation sheet.

    Two-level grouping: outer = cluster (matches the ``examples``
    sheet), inner = individual record (one block per cluster member).
    Within a record block, one row per field shows the field name +
    value at each of the 4 levels — so the reader can scan left-to-right
    and watch K1/K5/K6/K8/K10 transformations on a specific record.
    """
    ws = wb.create_sheet("transformations")
    headers = ["cluster", "record_id", "source", "field"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    wrap = Alignment(wrap_text=True, vertical="top")
    row_idx = 2
    last_cluster: int | None = None
    last_record: tuple[int, str] | None = None
    for block in transformations:
        cid = block["cluster_id"]
        rid = block["record_id"]
        if cid != last_cluster:
            # Outer (cluster) header.
            label = f"[{block['cluster_label']}]"
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col, value=label if col == 1 else "")
                cell.fill = _GROUP_FILL
                cell.font = _GROUP_FONT
            row_idx += 1
            last_cluster = cid

        if last_record != (cid, rid):
            # Inner (record) header — lighter visual emphasis: same fill
            # but italic text and the source/id labels in cols 2-3.
            sub_label = f"  {rid} ({block['source']})"
            for col in range(1, len(headers) + 1):
                cell = ws.cell(
                    row=row_idx, column=col, value=sub_label if col == 1 else ""
                )
                cell.fill = _GROUP_FILL
                cell.font = Font(italic=True)
            row_idx += 1
            last_record = (cid, rid)

        for baseline_name, per_level in block["fields"]:
            ws.cell(row=row_idx, column=1, value=cid)
            ws.cell(row=row_idx, column=2, value=rid)
            ws.cell(row=row_idx, column=3, value=block["source"])
            ws.cell(row=row_idx, column=4, value=baseline_name)
            for col_offset, level in enumerate(LEVELS, start=5):
                val = per_level.get(level, "")
                cell = ws.cell(
                    row=row_idx,
                    column=col_offset,
                    value=val if val != "" else None,
                )
                cell.alignment = wrap
            row_idx += 1

    _autosize(ws)
    # Widen the value columns so multi-token strings stay readable.
    for c in range(5, 5 + len(LEVELS)):
        ws.column_dimensions[get_column_letter(c)].width = 45


def _write_committee_summary_sheet(
    wb: Workbook,
    per_level_metrics: Mapping[str, Mapping[str, Any]],
) -> None:
    """Emit the committee summary sheet."""
    ws = wb.create_sheet("committee_summary")
    headers = ["stage", "metric"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    stages = ("sm", "norm", "em_blocking", "em_matching", "fusion")
    for stage in stages:
        # Row 1: committee aggregated metric.
        ws.cell(row=row_idx, column=1, value=stage)
        ws.cell(row=row_idx, column=2, value=f"committee_{_STAGE_AGG_KEY[stage]}")
        for col_offset, level in enumerate(LEVELS, start=3):
            metrics = per_level_metrics.get(level)
            if metrics is None:
                continue
            per_stage = metrics.get("per_stage") or {}
            stage_block = per_stage.get(stage) or {}
            v = _stage_agg_value(stage_block, stage)
            ws.cell(row=row_idx, column=col_offset, value=v)
        row_idx += 1

        # Row 2: best-member value.
        ws.cell(row=row_idx, column=1, value=stage)
        ws.cell(row=row_idx, column=2, value=f"best_member_{_STAGE_MEMBER_KEY[stage]}")
        for col_offset, level in enumerate(LEVELS, start=3):
            metrics = per_level_metrics.get(level)
            if metrics is None:
                continue
            per_stage = metrics.get("per_stage") or {}
            stage_block = per_stage.get(stage) or {}
            _, value = _stage_best_member(stage_block, stage)
            ws.cell(row=row_idx, column=col_offset, value=value)
        row_idx += 1

        # Row 3: best-member name.
        ws.cell(row=row_idx, column=1, value=stage)
        ws.cell(row=row_idx, column=2, value="best_member_name")
        for col_offset, level in enumerate(LEVELS, start=3):
            metrics = per_level_metrics.get(level)
            if metrics is None:
                continue
            per_stage = metrics.get("per_stage") or {}
            stage_block = per_stage.get(stage) or {}
            name, _ = _stage_best_member(stage_block, stage)
            ws.cell(row=row_idx, column=col_offset, value=name or None)
        row_idx += 1

    # Number-format the score cells.
    for r in ws.iter_rows(min_row=2, min_col=3, max_col=len(headers)):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    _autosize(ws)


def _write_selection_map_sheet(
    wb: Workbook,
    per_level_metrics: Mapping[str, Mapping[str, Any]],
) -> None:
    """Emit a per-level selection map for C12 optimized members.

    Reads ``notes.selection_map`` on each per-member block — populated by
    ``C12FusionCommitteeRunner`` (for ``pydi_per_attribute_optimal``) and
    ``C12NormCommitteeRunner`` (for ``rule_per_attribute_optimal``) — and
    writes one row per (stage, member, attribute) showing the picked
    method per level. Empty when no C12 optimized members are present.
    """
    ws = wb.create_sheet("selection_map")
    headers = ["stage", "member", "attribute"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    stages = ("norm", "fusion")
    for stage in stages:
        members = _all_members_for_stage(per_level_metrics, stage)
        if not members:
            continue
        # Build {member: {level: selection_map}} once per stage so we can
        # iterate (member, attribute) cleanly.
        per_member_maps: dict[str, dict[str, dict[str, str]]] = {}
        for member in members:
            level_to_map: dict[str, dict[str, str]] = {}
            for level in LEVELS:
                metrics = per_level_metrics.get(level)
                if metrics is None:
                    continue
                per_stage = metrics.get("per_stage") or {}
                stage_block = per_stage.get(stage) or {}
                per_member = stage_block.get("per_member") or {}
                body = per_member.get(member) or {}
                notes = body.get("notes") if isinstance(body, Mapping) else None
                if not isinstance(notes, Mapping):
                    continue
                smap = notes.get("selection_map")
                if isinstance(smap, Mapping) and smap:
                    level_to_map[level] = {str(k): str(v) for k, v in smap.items()}
            if level_to_map:
                per_member_maps[member] = level_to_map

        if not per_member_maps:
            continue

        # Group header row.
        for col in range(1, len(headers) + 1):
            cell = ws.cell(
                row=row_idx,
                column=col,
                value=f"[{stage}]" if col == 1 else "",
            )
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT
        row_idx += 1

        for member, level_to_map in per_member_maps.items():
            # Union of attributes across all levels for this member.
            attributes: set[str] = set()
            for smap in level_to_map.values():
                attributes.update(smap.keys())
            for attribute in sorted(attributes):
                ws.cell(row=row_idx, column=1, value=stage)
                ws.cell(row=row_idx, column=2, value=member)
                ws.cell(row=row_idx, column=3, value=attribute)
                for col_offset, level in enumerate(LEVELS, start=4):
                    smap = level_to_map.get(level) or {}
                    pick = smap.get(attribute)
                    ws.cell(row=row_idx, column=col_offset, value=pick or None)
                row_idx += 1

    _autosize(ws)


def _write_per_member_sheet(
    wb: Workbook,
    per_level_metrics: Mapping[str, Mapping[str, Any]],
) -> None:
    """Emit the detailed per-member sheet."""
    ws = wb.create_sheet("per_member")
    headers = ["stage", "member", "metric"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    stages = ("sm", "norm", "em_blocking", "em_matching", "fusion")
    for stage in stages:
        members = _all_members_for_stage(per_level_metrics, stage)
        if not members:
            continue
        # Full-width group header row (no merges — see sizes-sheet note).
        # Avoid a leading ``=`` (openpyxl auto-detects formulas).
        for col in range(1, len(headers) + 1):
            cell = ws.cell(
                row=row_idx, column=col, value=f"[{stage}]" if col == 1 else ""
            )
            cell.fill = _GROUP_FILL
            cell.font = _GROUP_FONT
        row_idx += 1

        member_key = _STAGE_MEMBER_KEY[stage]
        for member in members:
            ws.cell(row=row_idx, column=1, value=stage)
            ws.cell(row=row_idx, column=2, value=member)
            ws.cell(row=row_idx, column=3, value=member_key)
            for col_offset, level in enumerate(LEVELS, start=4):
                metrics = per_level_metrics.get(level)
                if metrics is None:
                    continue
                per_stage = metrics.get("per_stage") or {}
                stage_block = per_stage.get(stage) or {}
                per_member = stage_block.get("per_member") or {}
                body = per_member.get(member) or {}
                if not isinstance(body, Mapping):
                    continue
                member_metrics = body.get("metrics") or {}
                if not isinstance(member_metrics, Mapping):
                    continue
                v = member_metrics.get(member_key)
                try:
                    v_f = float(v) if v is not None else None
                except (TypeError, ValueError):
                    v_f = None
                ws.cell(row=row_idx, column=col_offset, value=v_f)
            row_idx += 1

    # Number-format the score cells.
    for r in ws.iter_rows(min_row=2, min_col=4, max_col=len(headers)):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    _autosize(ws)


def _write_evaluation_legend_sheet(wb: Workbook) -> None:
    """Emit the front-matter legend explaining stage evaluation surfaces.

    Written first (sheet index 0) so a reader sees it before any metric
    sheet. Documents the R7b dual-model dual-test infrastructure and
    points each surface at its dedicated sheet.
    """
    ws = wb.create_sheet("evaluation_legend", 0)

    rows: list[tuple[str, str]] = [
        ("Evaluation surfaces - what each sheet measures", ""),
        ("", ""),
        (
            "Why multiple surfaces exist",
            (
                "Under the R7b dual-model dual-test infrastructure "
                "(plan_revision.md R7b, 2026-05-27) every EM stage "
                "is evaluated on the cross-product of "
                "{baseline-trained model, variant-trained model} x "
                "{baseline test gold, variant-regenerated test gold} = "
                "4 separate evaluation surfaces. SM, Norm, and Fusion "
                "do not create additional test sets across variant "
                "levels (Fusion gold is copied from baseline at "
                "package time; SM / Norm have a single per-level gold) "
                "and they have no model train/test split — those "
                "stages report a single evaluation surface."
            ),
        ),
        ("", ""),
        ("EM matching sheets", "What each shows"),
    ]
    for surface in _EM_MATCHING_SURFACES:
        rows.append((surface["sheet_name"], surface["description"]))

    rows.extend(
        [
            ("", ""),
            ("EM blocking sheets", "What each shows"),
        ]
    )
    for surface in _EM_BLOCKING_SURFACES:
        rows.append((surface["sheet_name"], surface["description"]))

    rows.extend(
        [
            ("", ""),
            (
                "committee_summary - EM matching row",
                (
                    "The em_matching row in committee_summary is the "
                    "'train=Var test=Var' surface (paper headline). "
                    "The other three EM matching surfaces appear only "
                    "in their dedicated sheets above."
                ),
            ),
            ("", ""),
            (
                "committee_summary - EM blocking row",
                (
                    "The em_blocking row in committee_summary uses the "
                    "single-test macro_pair_recall (the only meaningful "
                    "blocking metric today; see runner note in the "
                    "EM block sheets)."
                ),
            ),
            ("", ""),
            (
                "Other stages (SM, Norm, Fusion)",
                (
                    "Single evaluation surface per level. SM macro_f1, "
                    "Norm macro_f1, and Fusion overall_accuracy are "
                    "unambiguous in committee_summary."
                ),
            ),
            ("", ""),
            (
                "Per-member runner state (R7c)",
                (
                    "Not every EM matching member actually retrains "
                    "under each variant. When the runner reuses the "
                    "same predictions for both models, the "
                    "'train=BL' and 'train=Var' sheets will show "
                    "identical values for that member. Divergence "
                    "between sheets surfaces only when R7c retrain "
                    "wiring is active for the member."
                ),
            ),
            ("", ""),
            (
                "EM blocking dual-test runner state",
                (
                    "Runner-side dual-test for blockers is not yet "
                    "wired (R7c follow-up). The four em_blocking "
                    "surface sheets are emitted forward-compatibly but "
                    "currently show 0.0 for every cell. They will "
                    "populate automatically once the runner produces "
                    "the dual-test values."
                ),
            ),
        ]
    )

    for r_idx, (col_a, col_b) in enumerate(rows, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=col_a)
        cell_b = ws.cell(row=r_idx, column=2, value=col_b)
        cell_a.alignment = Alignment(wrap_text=True, vertical="top")
        cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        if r_idx == 1:
            cell_a.font = _HEADER_FONT
        if r_idx == 5:
            cell_a.fill = _HEADER_FILL
            cell_a.font = _HEADER_FONT
            cell_b.fill = _HEADER_FILL
            cell_b.font = _HEADER_FONT

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 100


def _write_dual_test_surface_sheet(
    wb: Workbook,
    per_level_metrics: Mapping[str, Mapping[str, Any]],
    surface: Mapping[str, str],
) -> None:
    """Emit one dual-test surface sheet for ``surface['stage']``.

    Stage-agnostic — the surface dict carries the stage name, aggregated
    metric key, per-member metric key, and the human-readable
    sheet/title/description strings. Layout: title row + description
    row + spacer + header row + committee aggregated row + one row per
    individual member. Score cells are numeric, formatted to 4 decimals.
    """
    ws = wb.create_sheet(surface["sheet_name"])
    stage = surface["stage"]

    title_cell = ws.cell(row=1, column=1, value=surface["title"])
    title_cell.font = _HEADER_FONT
    desc_cell = ws.cell(row=2, column=1, value=surface["description"])
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Merge description across the value columns so wrapping is legible.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(LEVELS))

    headers = ["member", "metric"] + list(LEVELS)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 5
    # Committee aggregated row first.
    ws.cell(row=row_idx, column=1, value="<committee>")
    ws.cell(row=row_idx, column=2, value=surface["agg_key"])
    for col_offset, level in enumerate(LEVELS, start=3):
        metrics = per_level_metrics.get(level)
        if metrics is None:
            continue
        per_stage = metrics.get("per_stage") or {}
        stage_block = per_stage.get(stage) or {}
        agg = stage_block.get("aggregated") or {}
        v = agg.get(surface["agg_key"])
        try:
            v_f = float(v) if v is not None else None
        except (TypeError, ValueError):
            v_f = None
        ws.cell(row=row_idx, column=col_offset, value=v_f)
    row_idx += 1

    # Per-member rows in stable alpha order.
    members = _all_members_for_stage(per_level_metrics, stage)
    for member in members:
        ws.cell(row=row_idx, column=1, value=member)
        ws.cell(row=row_idx, column=2, value=surface["member_key"])
        for col_offset, level in enumerate(LEVELS, start=3):
            metrics = per_level_metrics.get(level)
            if metrics is None:
                continue
            per_stage = metrics.get("per_stage") or {}
            stage_block = per_stage.get(stage) or {}
            per_member = stage_block.get("per_member") or {}
            body = per_member.get(member) or {}
            mvals = body.get("metrics") or {}
            v = mvals.get(surface["member_key"])
            try:
                v_f = float(v) if v is not None else None
            except (TypeError, ValueError):
                v_f = None
            ws.cell(row=row_idx, column=col_offset, value=v_f)
        row_idx += 1

    for r in ws.iter_rows(min_row=5, min_col=3, max_col=len(headers)):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    _autosize(ws)


def build_workbook(domain: str, out_path: Path) -> Path:
    """Build and write the per-domain XLSX. Returns the written path."""
    per_level_metrics = _load_level_metrics(domain)
    if not per_level_metrics:
        raise FileNotFoundError(
            f"No metrics found for domain={domain!r}: "
            f"baseline + validation files are both missing."
        )

    sizes = _collect_sizes(domain)
    em_ordered, em_table = _collect_em_splits(domain)
    fusion_splits = _collect_fusion_splits(domain)
    examples = _collect_corner_examples(domain)
    transformations = _collect_record_transformations(domain)

    wb = Workbook()
    # Remove the default empty sheet.
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_sizes_sheet(wb, sizes)
    _write_splits_sheet(wb, em_ordered, em_table, fusion_splits)
    _write_examples_sheet(wb, examples)
    _write_transformations_sheet(wb, transformations)
    _write_committee_summary_sheet(wb, per_level_metrics)
    _write_per_member_sheet(wb, per_level_metrics)
    _write_selection_map_sheet(wb, per_level_metrics)
    for surface in _ALL_DUAL_TEST_SURFACES:
        _write_dual_test_surface_sheet(wb, per_level_metrics, surface)
    # Legend written last so its insert-at-position-0 lands it as the
    # leftmost (first) tab the reader sees on opening the workbook.
    _write_evaluation_legend_sheet(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv: list[str] | None = None) -> None:
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description=(
            "Build per-domain XLSX statistics under "
            "usecases_synthetic/statistics/. See module docstring for "
            "details."
        )
    )
    parser.add_argument(
        "--domain",
        action="append",
        default=None,
        help=(
            "Domain name (may be repeated). When omitted, builds for "
            "every domain in the default set with at least a baseline "
            f"present: {', '.join(DEFAULT_DOMAINS)}."
        ),
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=STATISTICS_DIR,
        help=("Output directory. Defaults to " "usecases_synthetic/statistics/."),
    )
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    domains = args.domain or list(DEFAULT_DOMAINS)
    written: list[Path] = []
    for d in domains:
        out_path = args.out_dir / f"{d}.xlsx"
        try:
            written.append(build_workbook(d, out_path))
            logger.info("Wrote %s", out_path)
        except FileNotFoundError as exc:
            logger.warning("Skipping %s: %s", d, exc)

    if not written:
        logger.error("No statistics workbooks were written.")
        sys.exit(1)

    print(f"Wrote {len(written)} workbook(s) under {args.out_dir}:")
    for p in written:
        try:
            rel = p.relative_to(REPO_ROOT)
        except ValueError:
            rel = p
        print(f"  - {rel}")


if __name__ == "__main__":
    main()
